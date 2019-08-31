"""
Description: Simple Python script that reads through a sample excel file and prints it.
Module: openpyxl is a Python library to read/write Excel 2010 xlsx/xlsm/xltx/xltm files.
Install: pip install openpyxl
"""
from openpyxl import load_workbook

source = 'samples/censuspopdata.xlsx'

# load excel file into a variable
wb = load_workbook(source)

# gets the workbooks active sheet
sheet = wb.active

# read through the censuspopdata.xlsx
for i in range(1, (sheet.max_row + 1)):
    for j in range(1, sheet.max_column + 1):
        print(sheet.cell(row=i, column=j).value, end=' ')
    print('\n', end='')
