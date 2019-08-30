"""
Description: Simple Python script that reads through a sample excel file and prints it.
Module: openpyxl is a Python library to read/write Excel 2010 xlsx/xlsm/xltx/xltm files.
Install: pip install openpyxl
"""
from openpyxl import load_workbook

source = 'samples/example.xlsx'

# load excel file into a variable
wb = load_workbook(source)

# gets the workbooks active sheet
sheet = wb.active

# read through the example.xlsx
for i in range(1, (sheet.max_row + 1)):
    for j in range(1, sheet.max_column + 1):
        print(sheet.cell(row=i, column=j).value, end=' ')
    print('\n', end='')

"""
Output:

2015-04-05 13:34:02 Apples 73 
2015-04-05 03:41:23 Cherries 85 
2015-04-06 12:46:51 Pears 14 
2015-04-08 08:59:43 Oranges 52 
2015-04-10 02:07:00 Apples 152 
2015-04-10 18:10:37 Bananas 23 
2015-04-10 02:40:46 Strawberries 98
"""